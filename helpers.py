import os
from matplotlib import pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from pandas import DataFrame
import seaborn as sns
from config import PERSON_VAT_COL, TYPE_PERSON_COL, output_path

from config import PLOTS_PATH, RESULT_COL, SUM


def generate_categories(df: DataFrame, output_path):
    generate_category_plot(TYPE_PERSON_COL, SUM, df, "C2", output_path)
    generate_count_plot(TYPE_PERSON_COL, df, output_path)
    generate_category_plot(RESULT_COL, SUM, df, "B2", output_path)
    generate_count_plot(RESULT_COL, df, output_path)


def generate_count_plot(column1, df, output_path):
    counts = df[column1].value_counts(dropna=False).reset_index()
    counts.columns = [column1, "Count"]
    counts["Percent"] = counts["Count"] / counts["Count"].sum() * 100

    y_labels = counts[column1].astype(str)

    plt.figure(figsize=(10, 6))
    ax = plt.gca()
    bars = ax.barh(y_labels, counts["Count"])

    total = counts["Count"].sum()
    labels = [f"{c} ({c/total:.1%})" for c in counts["Count"]]
    ax.bar_label(bars, labels=labels, label_type="edge")

    ax.set_title(f"Брой плащания по {column1}")
    ax.set_xlabel("Брой")
    ax.set_ylabel(f"Тип {column1}")
    plt.tight_layout()

    plot_path = os.path.join(PLOTS_PATH, f"boxplot{column1}_Count.png")
    plt.savefig(plot_path, bbox_inches="tight")
    plt.close()

    insert_plot(plot_path, output_path, column1)


def generate_category_plot(column1, column2, df: DataFrame, cell, output_path):
    plt.figure(figsize=(20, 6))
    sns.boxplot(x=column1, y=column2, data=df)
    mean = df[SUM].mean()
    plt.ylim(0, mean)
    plt.title(f"{column1} по {column2}")

    plot_path = os.path.join(PLOTS_PATH, f"boxplot{column1}-{column2}.png")
    plt.savefig(plot_path, bbox_inches="tight")
    plt.close()

    insert_plot(plot_path, output_path, column1, column2)


def insert_plot(plot_path, output_path, column1, column2=""):
    wb = load_workbook(output_path)
    ws = wb.create_sheet(f"{column1} with {column2}")

    img = Image(plot_path)
    ws.add_image(img, "B2")

    wb.save(output_path)
